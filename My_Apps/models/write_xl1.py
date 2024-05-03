import imp
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import webbrowser
from flask import send_file,send_from_directory
import os
import time


### get data from master sheet
class write_xl1:
    # def __init__(self,names,type,object_id,device_id,object_name,read_write,unit,min1,max1,normal_state,desc):
    def __init__(self,names,read_write,unit,min1,max1,normal_state,desc):
        self.names=names
        # self.type=type
        # self.object_id=object_id
        # self.device_id=device_id
        # self.object_name=object_name
        self.read_write=read_write
        self.unit=unit
        self.min=min1
        self.max=max1
        self.normal_state=normal_state
        self.desc=desc
        # self.plzout(self.names,self.type,self.object_id,self.device_id,self.object_name,self.read_write,self.unit,self.min,self.max,self.normal_state,self.desc)
        self.plzout()

    def plzout(self):
        wb=openpyxl.Workbook()
        wb['Sheet'].title="point_list"
        sheet = wb.active
################################ enter header with two ways ######
        ### WAY 1:
        c1=sheet.cell(row=1,column=3)
        c1.value="Name"
        c1.fill=PatternFill("solid",fgColor="71FF33") ### STYLE THE CELL
        # c1a=sheet.cell(row=1,column=4)
        # c1a.value="Type"
        # c1a.fill=PatternFill("solid",fgColor="71FF33") ### STYLE THE CELL
        # c2=sheet.cell(row=1,column=5)
        # c2.value="Object ID"
        # c2.fill=PatternFill("solid",fgColor="71FF33")
        # c2a=sheet.cell(row=1,column=6)
        # c2a.value="Device ID"
        # c2a.fill=PatternFill("solid",fgColor="71FF33")
        
        #### WAY2:
        # sheet['G1'].value="Object Name"
        # sheet['G1'].fill=PatternFill("solid",fgColor="71FF33")
        sheet['H1'].value="Read/Write"
        sheet['H1'].fill=PatternFill("solid",fgColor="71FF33")
        c5=sheet.cell(row=1,column=9)
        c5.value="Unit"
        c5.fill=PatternFill("solid",fgColor="71FF33")
        c6=sheet.cell(row=1,column=10)
        c6.value="Min."
        c6.fill=PatternFill("solid",fgColor="71FF33")
        c7=sheet.cell(row=1,column=11)
        c7.value="Max."
        c7.fill=PatternFill("solid",fgColor="71FF33")
        c8=sheet.cell(row=1,column=12)
        c8.value="Normal Status"
        c8.fill=PatternFill("solid",fgColor="71FF33")
        c9=sheet.cell(row=1,column=13)
        c9.value="Description"
        c9.fill=PatternFill("solid",fgColor="71FF33")
      
        a1=0
        for t1 in range (2,(len(self.names)+2)):
            z1=sheet.cell(row=t1,column=3)
            z1.value=self.names[a1]
            a1+=1

        # a2=0
        # for t2 in range (2,(len(self.type)+2)):
        #     z2=sheet.cell(row=t2,column=4)
        #     z2.value=self.type[a2]
        #     a2+=1

        # a3=0
        # for t3 in range (2,(len(self.object_id)+2)):
        #     z3=sheet.cell(row=t3,column=5)
        #     z3.value=self.object_id[a3]
        #     a3+=1

        # a4=0
        # for t4 in range (2,(len(self.device_id)+2)):
        #     z4=sheet.cell(row=t4,column=6)
        #     z4.value=self.device_id[a4]
        #     a4+=1

        # a5=0
        # for t5 in range (2,(len(self.object_name)+2)):
        #     z5=sheet.cell(row=t5,column=7)
        #     z5.value=self.object_name[a5]
        #     a5+=1

        a6=0
        for t6 in range (2,(len(self.read_write)+2)):
            z6=sheet.cell(row=t6,column=8)
            z6.value=self.read_write[a6]
            a6+=1

        a7=0
        for t7 in range (2,(len(self.unit)+2)):
            z7=sheet.cell(row=t7,column=9)
            z7.value=self.unit[a7]
            a7+=1

        a8=0
        for t8 in range (2,(len(self.min)+2)):
            z8=sheet.cell(row=t8,column=10)
            z8.value=self.min[a8]
            a8+=1

        a9=0
        for t9 in range (2,(len(self.max)+2)):
            z9=sheet.cell(row=t9,column=11)
            z9.value=self.max[a9]
            a9+=1

        a10=0
        for t10 in range (2,(len(self.normal_state)+2)):
            z10=sheet.cell(row=t10,column=12)
            z10.value=self.normal_state[a10]
            a10+=1

        a11=0
        for t11 in range (2,(len(self.desc)+2)):
            z11=sheet.cell(row=t11,column=13)
            z11.value=self.desc[a11]
            a11+=1
        wb.save("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx")
        # time.sleep(10)
        
        ### DOWNLOAD THE FILE
        
        
        return send_file("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",as_attachment=True)
        
        
        # chrome_path="C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
        # webbrowser.register('chrome', None,webbrowser.BackgroundBrowser(chrome_path))   
        # browser='chrome'

        # # #### open new chrome tab
        # webbrowser.get(browser).open_new_tab("X:/M.HAMMAD/Output sheet.xlsx")
        




if __name__=='__main__':
    names=input("enter name"+" ")
    type=input("enter add"+" ")
    object_id=input("enter object_id"+" ")
    device_id=input("enter device_id"+" ")
    object_name=input("enter object_name"+" ")
    read_write=input("enter read_write"+" ")
    unit=input("enter unit"+" ")
    min=input("enter min"+" ")
    max=input("enter max"+" ")
    normal_state=input("enter normal_state"+" ")
    desc=input("enter desc"+" ")
    
    
 

    wtx=write_xl1(names,type,object_id,device_id,object_name,read_write,unit,min,max,normal_state,desc)
                                                                             
       

    



