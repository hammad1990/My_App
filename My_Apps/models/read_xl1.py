from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



### get data from master sheet
class read_xls1:
    def __init__(self,path1,path2):
        self.path1=path1
        self.path2=path2
        workbook = load_workbook(filename=path1)
        workbook1 = load_workbook(filename=path2)
        self.sheet1 = workbook.active
        self.sheet2 = workbook1.active
        # print("dasda")
###############################        
        self.final_names=[]
        self.final_type=[]
        self.final_object_id=[]
        self.final_device_id=[]
        self.final_object_name=[]
        self.final_read_write=[]
        self.final_unit=[]
        self.final_min=[]
        self.final_max=[]
        self.final_normal_state=[]
        self.final_desc=[]
###############################         
        
        self.mst_points=[]
        self.read_write=[]
        self.unit=[]
        self.min=[]
        self.max=[]
        self.normal_state=[]
        self.desc=[]
###############################
        self.pls_points=[]
        self.type=[]
        self.object_id=[]
        self.device_id=[]
        self.object_name=[]
        
        self.get_master()
        self.get_alaa()
    def get_master(self):   
        for i1 in range(2,self.sheet1.max_row+1):  ## start from row 2
            for j1 in range(3,4):   ##get point names
                cell1=self.sheet1.cell(row=i1,column=j1)
                self.mst_points.append(cell1.value)
            for j1h in range(8,9):   ### get Read/Write
                cell1h=self.sheet1.cell(row=i1,column=j1h)
                self.read_write.append(cell1h.value)
            for j1c in range(9,10):   ### get unit
                cell1c=self.sheet1.cell(row=i1,column=j1c)
                self.unit.append(cell1c.value)
            for j1a in range(10,11):   ### get min
                cell1a=self.sheet1.cell(row=i1,column=j1a)
                self.min.append(cell1a.value)
            for j1b in range(11,12):   ### get max
                cell1b=self.sheet1.cell(row=i1,column=j1b)
                self.max.append(cell1b.value)
            for j1i in range(12,13):   ### get Normal Status
                cell1i=self.sheet1.cell(row=i1,column=j1i)
                self.normal_state.append(cell1i.value)
            for j1d in range(13,14):   ### get desc
                cell1d=self.sheet1.cell(row=i1,column=j1d)
                self.desc.append(cell1d.value)
            
##        print(self.mst_points)            
      
        ### get data from point list sheet
    def get_alaa(self):
                
        for i2 in range(1,self.sheet2.max_row+1):## start from row 1
            for j2 in range(1,2):       ##get names
                cell2=self.sheet2.cell(row=i2,column=j2)
                if cell2.value==None:
                    pass
                else:
                    self.pls_points.append(cell2.value)

        self.final()
            
                       
    def final(self):       
                    
        for x in range (0,len(self.mst_points)):
            for y in range (0,len(self.pls_points)):
                if self.mst_points[x]==self.pls_points[y]:
                    self.final_names.append(self.mst_points[x])
                    self.final_read_write.append(self.read_write[x])
                    self.final_unit.append(self.unit[x])
                    self.final_min.append(self.min[x])
                    self.final_max.append(self.max[x])
                    self.final_normal_state.append(self.normal_state[x])
                    self.final_desc.append(self.desc[x])
                    
                    
                    break


        return (self.final_names,self.final_read_write,self.final_unit,self.final_min,self.final_max,self.final_normal_state,self.final_desc)

if __name__=='__main__':
    p1="general point list.xlsx"
    p2="P212556-morgan_lewis_-_2222_market_st_point list.xlsx"
    rtx=read_xls1(p1,p2)
    



