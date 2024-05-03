# from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import PIL
from datetime import datetime,date,timedelta,time
from flask import send_file,send_from_directory
from openpyxl.styles import Alignment



class Create_excel_component:
  def __init__(self,co_project_no,co_item_model,co_item_qty,co_item_no,co_item_ref,co_ambient,co_voltage,compressors,condensers,supply):
    self.co_project_no=co_project_no
    self.co_item_model=co_item_model
    self.co_item_qty=co_item_qty
    self.co_item_no=co_item_no
    self.co_item_ref=co_item_ref
    self.co_ambient=co_ambient
    self.co_voltage=co_voltage
    self.compressors=compressors
    self.condensers=condensers
    self.supply=supply
    self.current_date = datetime.today().date()
    
    
    self.create_component()
    # print(self.supply)
  def create_component(self):
   
    
    wb = load_workbook('Z:/M.HAMMAD/PY/My_Apps/Excel files/template.xlsx')
    sheet = wb.active


    sheet['D6'].value=self.co_project_no
    sheet['D7'].value=self.co_item_model
    sheet['D8'].value=self.co_item_qty
    sheet['D9'].value=self.co_item_no
    sheet['D10'].value=self.co_item_ref
    sheet['O6'].value=f"USA/{self.co_ambient} °F"
    sheet['O7'].value=self.co_voltage
    sheet['O8'].value=self.current_date
    sheet['B21'].alignment =Alignment(horizontal='center')
    sheet['B21'].value=f"R-410A//{len(self.compressors)} INDEPENDENT CIRCUITS-EXV FOR ALL COMPRESSORS"

      
###################### print the compressors###################
  
    x=1
    for count,value in enumerate(self.compressors.values(),start=15):

      if value[1]=="VZH035CJANB"  or value[1]=="VZH044CJANB/M" or value[1]=="VZH065CJANB" or value[1]=="VZH088BJANA/I/P06" or value[1]=="VZH088CJAN" or value[1]=="VZH088AGANA" or value[1]=="VZH117AGANA": 
        sheet['B'+str(count)]=f"{value[0]}-{value[1]} ({value[3]}HP)-CIR.{value[2]}"                      #DANFOSS-VZH117AGANA ((15 HP))---CIR.1
        sheet['F'+str(count)]=f"{value[4]}A/{value[5]}A"                                                                     #RLA/PT
        sheet['H'+str(count)] = f"Cable inside ELEC.Panel:{value[6]}AWG/Cable outside ELEC.Panel:{value[7]}AWG"  #CABLE IN AND OUT
        sheet['I'+str(count)] =f"{value[6]}AWG"                                                                   #GROUND CABLE
        sheet['K'+str(count)] =f"{value[8]}"                                                                      #INVERTER
        sheet['O'+str(count)] =f"{value[9]}"                                                                       #BREAKER
        
            
      else:
        sheet['B'+str(count)]=f"{value[0]}-{value[1]} ({value[3]}HP)-CIR.{value[2]}"   #COPELAND-ZP61-5HP--CIR-1
        sheet['F'+str(count)]=f"{value[4]}A/{value[5]}A"                                          #RLA/PT
        sheet['H'+str(count)] =f"{value[6]}AWG"                                      #CABLE
        sheet['I'+str(count)] =f"{value[6]}AWG"                                       #GROUND CABLE
        sheet['K'+str(count)] =f"{value[7]}"                                       #CONTACTOR
        sheet['O'+str(count)] =f"{value[8]}"                                       #MMS/breaker
        
        
      
###################### print the condenser#####################################################
    
    if  self.condensers['Brand']!="NOTHING":   
      sheet['B27']=f"{self.condensers['Brand']}-{self.condensers['model']}"   #EBM-S6D910AB0505-DB
      sheet['D27']=f"{self.condensers['HP']}"    #HP
      sheet['E27']=f"{self.condensers['qty']}"   #QTY
      sheet['F27']=f"{self.condensers['FLA']}A"   #FLA
      sheet['G27']=f"{self.condensers['Cable']}" #Cable
      sheet['H27']=f"{self.condensers['Cable']}" #Cable

      ####### case 1####################################
      if int(self.condensers['qty']) <=3:
        sheet['I27']=f"{self.condensers['Contactor']}" #Contactor
        sheet['K27']=self.condensers['qty'] #QTY OF CONTACTORS
        sheet['P27']=f"{self.condensers['qty']}*{self.condensers['Breaker']}" #QTY & BREAKERS
      ####### case 2####################################
      elif int(self.condensers['qty']) <=3  and self.condensers['speed']=="Speed-10A-UL-ZIEHL-PKDM10":
        sheet['I27']=f"{self.condensers['Contactor']}" #Contactor
        sheet['K27']=int(self.condensers['qty']) #QTY OF CONTACTORS
        sheet['P27']=f"{int(self.condensers['qty'])}*{self.condensers['Breaker']}" #QTY & BREAKERS
        sheet['B28']=f"{self.condensers['speed']}-QTY=1" #QTY of speed
      ####### case 3####################################
      elif int(self.condensers['qty']) <=3  and self.condensers['speed']!=" " and self.condensers['speed']!="Speed-10A-UL-ZIEHL-PKDM10" :
        sheet['I27']=f"{self.condensers['Contactor']}" #Contactor
        sheet['K27']=f"{int(self.condensers['qty'])-1}" #QTY OF CONTACTORS
        sheet['P27']=f"{int(self.condensers['qty'])}*{self.condensers['Breaker']}" #QTY & BREAKERS
        sheet['B28']=f"{self.condensers['speed']}-QTY=1" #QTY of speed
      ####### case 4####################################
      elif int(self.condensers['qty']) ==6  and self.condensers['speed']==" ":
        if self.condensers['model']=="ZC080-SDL.6N.V7" or self.condensers['model']=="S6D910AB0505-DB" or self.co_voltage=="S6D800-AE05-09"\
        and self.co_voltage=="208V-3-60Hz" or self.co_voltage=="220V-3-60Hz" or self.co_voltage=="230V-3-60Hz" :
          sheet['I27']="ABB-AF30A" #Contactor
        sheet['K27']=int(self.condensers['qty'])-3 #QTY OF CONTACTORS
        sheet['P27']=f"{int(self.condensers['qty'])-3}*ABB-32A" #QTY & BREAKERS
      ####### case 5####################################
      elif int(self.condensers['qty']) ==6  and self.condensers['speed']=="Speed-10A-UL-ZIEHL-PKDM10" and  self.condensers['vee']!="2Vee":
        if self.condensers['model']=="ZC080-SDL.6N.V7" or self.condensers['model']=="S6D910AB0505-DB" or self.co_voltage=="S6D800-AE05-09"\
        and self.co_voltage=="208V-3-60Hz" or self.co_voltage=="220V-3-60Hz" or self.co_voltage=="230V-3-60Hz" :
          sheet['I27']="ABB-AF30A" #Contactor
          sheet['P27']=f"{int(self.condensers['qty'])-3}*ABB-32A" #QTY & BREAKERS
        sheet['K27']=int(self.condensers['qty'])-3 #QTY OF CONTACTORS
        sheet['B28']=f"{self.condensers['speed']}-QTY=2" #QTY of speed
      ####### case 6####################################
      elif int(self.condensers['qty']) ==6  and self.condensers['speed']=="Speed-10A-UL-ZIEHL-PKDM10" and  self.condensers['vee']=="2Vee":
        if self.condensers['model']=="ZC080-SDL.6N.V7" or self.condensers['model']=="S6D910AB0505-DB" or self.co_voltage=="S6D800-AE05-09"\
        and self.co_voltage=="208V-3-60Hz" or self.co_voltage=="220V-3-60Hz" or self.co_voltage=="230V-3-60Hz" :
          sheet['I27']="ABB-AF30A" #Contactor
          sheet['P27']=f"{int(self.condensers['qty'])-3}*ABB-32A" #QTY & BREAKERS
        sheet['K27']=int(self.condensers['qty'])-3 #QTY OF CONTACTORS
        sheet['B28']=f"{self.condensers['speed']}-QTY=4" #QTY of speed
      ####### case 7####################################
      elif int(self.condensers['qty']) ==6 and self.condensers['speed']!=" "and self.condensers['speed']!="Speed-10A-UL-ZIEHL-PKDM10":
        if self.condensers['model']=="ZC080-SDL.6N.V7" or self.condensers['model']=="S6D910AB0505-DB" or self.co_voltage=="S6D800-AE05-09"\
        and self.co_voltage=="208V-3-60Hz" or self.co_voltage=="220V-3-60Hz" or self.co_voltage=="230V-3-60Hz" :
          sheet['I27']="ABB-AF30A" #Contactor
          sheet['P27']=f"2*ABB-32A+2*ABB-16A" #QTY & BREAKERS
        sheet['K27']=int(self.condensers['qty'])-4 #QTY OF CONTACTORS
        sheet['B28']=f"{self.condensers['speed']}-QTY=2" #QTY of speed
      ####### case 8####################################
      elif int(self.condensers['qty']) ==12  and self.condensers['speed']==" ":
        if self.condensers['model']=="ZC080-SDL.6N.V7" or self.condensers['model']=="S6D910AB0505-DB" or self.co_voltage=="S6D800-AE05-09"\
        and self.co_voltage=="208V-3-60Hz" or self.co_voltage=="220V-3-60Hz" or self.co_voltage=="230V-3-60Hz" :
          sheet['I27']="ABB-AF30A" #Contactor
          sheet['P27']=f"{int(self.condensers['qty'])-6}*ABB-32A" #QTY & BREAKERS
        sheet['K27']=int(self.condensers['qty'])-6 #QTY OF CONTACTORS
      ####### case 9####################################
      elif int(self.condensers['qty']) ==12  and self.condensers['speed']=="Speed-10A-UL-ZIEHL-PKDM10" :
        if self.condensers['model']=="ZC080-SDL.6N.V7" or self.condensers['model']=="S6D910AB0505-DB" or self.co_voltage=="S6D800-AE05-09"\
        and self.co_voltage=="208V-3-60Hz" or self.co_voltage=="220V-3-60Hz" or self.co_voltage=="230V-3-60Hz" :
          sheet['I27']="ABB-AF30A" #Contactor
          sheet['P27']=f"{int(self.condensers['qty'])-6}*ABB-32A" #QTY & BREAKERS
        sheet['K27']=int(self.condensers['qty'])-6 #QTY OF CONTACTORS
        sheet['B28']=f"{self.condensers['speed']}-QTY=2" #QTY of speed
      ####### case 10####################################
      elif int(self.condensers['qty']) ==12 and self.condensers['speed']!=" "and self.condensers['speed']!="Speed-10A-UL-ZIEHL-PKDM10":
        if self.condensers['model']=="ZC080-SDL.6N.V7" or self.condensers['model']=="S6D910AB0505-DB" or self.co_voltage=="S6D800-AE05-09"\
        and self.co_voltage=="208V-3-60Hz" or self.co_voltage=="220V-3-60Hz" or self.co_voltage=="230V-3-60Hz" :
          sheet['I27']="ABB-AF30A" #Contactor
          sheet['P27']=f"{int(self.condensers['qty'])-4}*ABB-32A" #QTY & BREAKERS
        sheet['K27']=int(self.condensers['qty'])-8 #QTY OF CONTACTORS
        sheet['B28']=f"{self.condensers['speed']}-QTY=4" #QTY of speed

    else:
      pass
    
    # print(self.condensers['speed'])
    # print(self.condensers['vee'])

###################### print the supply#####################################################

    for count,type in enumerate( self.supply,start=34):
      # print(type)
      sheet['B'+str(count)]=type                         #supply.return.E,wheel
      for count,value in enumerate(self.supply.values(),start=34):

        # print(type)
        
        sheet['C'+str(count)]=f"{value[0]}"                #MODEL
        sheet['D'+str(count)] = f"{value[1]}"              #HP
        sheet['E'+str(count)] = f"{value[2]}"              #QTY
        sheet['F'+str(count)] =f"{value[3]}"               #RLA
        sheet['G'+str(count)] =f"{value[4]}"               #CABLE
        sheet['H'+str(count)] =f"{value[4]}"               #PE CABLE
        sheet['O'+str(count)] =f"{value[4]}"               #MAIN CABLE
        sheet['I'+str(count)] =f"{value[2]}*{value[5]}"               #INVERTER
        sheet['P'+str(count)] =f"{value[2]}*{value[6]}"               #BREAKER


###################### print the options#####################################################
    sheet['B53']=f"**Flood detector--------HDSP------------NDPL--------PFR=SYMCOM--EVD-TWIN TYPE[RVLV0102039]-QTY={int(len(self.compressors)/2)}----SUCTION PRESS TRANSM-CAREL-SPKT0043P0\
      [RECN0137208]+CABLE[RVLV0105002]-QTY={int(len(self.compressors))}"
    sheet['B56']=f"*DISCHARGE & LIQUID TEMP SENSORS-MAMAC-PTX-704-A-12-B-[RECN5001152]--QTY={int(len(self.compressors)*2)}" #DISC+liquid TEMP SENSORS
    sheet['B57']=f"*LIQUID PRESS TRANSM-MCS-667C-20M[RECN0137220]--QTY={len(self.compressors)}" #liquid PRESS TRANM
    sheet['B58']=f"*SUCTION PRESS TRANSM-MCS-500C-20M[RECN0137219]--QTY={len(self.compressors)}" #SUCTION PRESS TRANM





    LOC1=f"Z:/M.hammad/components output/{self.co_project_no} {self.co_item_model} item{self.co_item_no}.xlsx"
    wb.save(LOC1)
      
    ### DOWNLOAD THE FILE
    
    
    # return send_file(LOC1,as_attachment=True)
    # return(LOC1)
    
      
      




# if __name__=='__main__':
  # project_name=input("enter project_name"+" ")
  # unit_name=input("enter unit_name"+" ")
  # names=input("enter name"+" ")
  # type=input("enter add"+" ")
  # object_id=input("enter object_id"+" ")
  # device_id=input("enter device_id"+" ")
  # object_name=input("enter object_name"+" ")
  # read_write=input("enter read_write"+" ")
  # unit=input("enter unit"+" ")
  # min=input("enter min"+" ")
  # max=input("enter max"+" ")
  # normal_state=input("enter normal_state"+" ")
  # desc=input("enter desc"+" ")
  
  


  # wtx=write_excel(project_name,unit_name,names,type,object_id,device_id,object_name,read_write,unit,min,max,normal_state,desc)
                                                                            
      

  


