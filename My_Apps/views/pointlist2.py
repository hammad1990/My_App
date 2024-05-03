from flask import Blueprint,render_template,request,session,redirect,url_for,session,send_file
from openpyxl import load_workbook
import openpyxl
from models.read_xl1 import read_xls1
from models.read_xl2 import read_xls2
from models.write_xl1 import write_xl1
from models.write_excel_R0 import write_excel
import os
import time
import pyodbc
list=[]

pointlist2=Blueprint("pointlist2",__name__,template_folder='templates',static_folder='static')
@pointlist2.route("/pointlist2", methods=["GET","POST"]) 
def pointlist2_func():

    if "user"in session:
        username=session["user"]
    else:
        return redirect('login')

    if request.method=="POST":
        Model=request.form['UModel']  ##pph/wpph
        DXQTY=request.form['DXQTY']   ## DX QTY
        HWCTSQTY=request.form['HWCTSQTY']   ## HWCTSQTY QTY
        CWCTSQTY=request.form['CWCTSQTY']   ## CWCTSQTY QTY
        HEATERQTY=request.form['HEATERQTY']   ## DX QTY
        COMPQTY=request.form['COMPQTY']   ## DX QTY
        GHQTY=request.form['GHQTY']   ## GH QTY
        SDQTY=request.form['SDQTY']   ## SD QTY
        DPSQTY=request.form['DPSQTY']   ## PRE FILTER QTY
        DPS3QTY=request.form['DPS3QTY']   ## FINAL FILTER QTY
        EFVFDQTY=request.form['EFVFDQTY']   ## EF VFD  QTY
        SFVFDQTY=request.form['SFVFDQTY']   ## SF VFD  QTY
        SRQTY=request.form['SRQTY']   ## SR  QTY
        options=request.form.getlist('fn1') #### options
        alaa_list= request.files['project file']
        general_list= request.files['general file']
  

        print(COMPQTY)
       
       

        # COMMON POINTS
        list.append("General Alarm Status")
        list.append("OAT SENSOR")
        list.append("SAT SENSOR")
        list.append("Unit Enable (BAS)")
        list.append("Control Source")
        list.append("Enable mode (BMS)")
        # list.append("Heat Setpoint")
        list.append("Cool Setpoint")
        list.append("COOLING SAT SETPOINT")
        list.append("COOLING Temp SETPOINT_Write")
        list.append("sat_stpt_ht_WRITE")
        list.append("Enable mode (BMS)_Write")
        list.append("Alarm Reset")


    #    compressor points
        for w in range (1,int(COMPQTY)+1):
            x="comp{}_liq_pr".format(w)
            list.append(x)
            x="comp{}_suc_pr".format(w)
            list.append(x)
            x="comp{}_SUC_pr".format(w)
            list.append(x)
            x="DA{} TEMP".format(w)
            print(x)
            list.append(x)
            x="C{}_HPS".format(w)
            list.append(x)
            x="C{}-HPS".format(w)
            list.append(x)
            x="C{}_HI".format(w)
            list.append(x)
            x="C{}_LPS".format(w)
            list.append(x)
            x="C{}_LPS".format(w)
            list.append(x)
            x="C{}-STATUS".format(w)
            list.append(x)
            x="C{}_MP".format(w)
            list.append(x)
            x="comp{}_flt".format(w)
            list.append(x)
            x="LIQ{} TEMP".format(w)
            list.append(x)
            print(x)
            
        #    ALL points
        for e in options:
            if e=='DX AIR TEMP':
                for i in range(1,(int(DXQTY)+1)) :
                    x="{}{}".format(e,i)
                    list.append(x)
            elif e=='HWCTS Sensor':
                for i in range(1,(int(HWCTSQTY)+1)) :
                    x="{}".format(e)
                    list.append(x)
                    x="{}{}".format(e,i)
                    list.append(x)
            
            elif e=='CWCTS Sensor':
                for i in range(1,(int(CWCTSQTY)+1)) :
                    x="{}".format("CWRT")
                    list.append(x)
                    x="{}".format("SCWRT")
                    list.append(x)
                    x="{}".format("CWST")
                    list.append(x)
                    x="{}".format("SCWST")
                    list.append(x)
                    x="{}{}".format("CWST",i)
                    list.append(x)
                    x="{}{}".format("SCWST",i)
                    list.append(x)
                    x="{}{}{}{}".format("2WV",i,"","OPEN%")
                    list.append(x)
                    
            
            elif e=='Heating Stage':
                for i in range(1,(int(HEATERQTY)+1)) :
                    x="{}{}{}{}{}".format(e," ",i," ",'Status')
                    list.append(x)
                    
            elif e=='GAS HEATER OUTPUT':
                for i in range(1,(int(GHQTY)+1)) :
                    x="{}{}{}{}".format("GAS HEATER",i," ",'OUTPUT')
                    list.append(x)
                    x="{}{}{}{}".format("GH",i," ",'FAULT')
                    list.append(x)
            elif e=='SD_flt1':
                for i in range(1,(int(SDQTY)+1)) :
                    x="{}{}".format("SD_flt",i)
                    list.append(x)
            
            elif e=='Pre Filter Change Required':
                for i in range(1,(int(DPSQTY)+1)) :
                    x="{}{}".format(e,i)
                    list.append(x)
            
            elif e=='Final Filter Change Required':
                for i in range(1,(int(DPS3QTY)+1)) :
                    x="{}{}".format(e,i)
                    list.append(x)

            elif e=='EF VFD %':
                for i in range(1,(int(EFVFDQTY)+1)) :
                    x="{}{}{}{}".format("EF",i,"","VFD %")
                    list.append(x)
                    x="{}{}{}{}".format("EF",i,"-","STATUS")
                    list.append(x)
                    x="{}{}{}{}".format("EX_FAN",i,"","Alarm Status")
                    list.append(x)
            elif e=='SF VFD %':
                for i in range(1,(int(SFVFDQTY)+1)) :
                    x="{}{}{}{}".format("SF",i," ","VFD %")
                    list.append(x)
                    x="{}{}{}{}".format("SF",i,"-","STATUS")
                    list.append(x)
                    x="{}{}{}{}".format("SUP_FAN",i," ","Alarm Status")
                    list.append(x)
            elif e=='fire alarm':
                x="{}".format(e)
                list.append(x)
            elif e=='Heat Setpoint':
                x="{}".format(e)
                list.append(x)
            elif e=='ZRH_R':
                x="{}".format(e)
                list.append(x)
                x="{}".format("ZTEMP_R")
                list.append(x)
            elif e=='Flood_al':
                x="{}".format(e)
                list.append(x)
            elif e=='NDPL_al':
                x="{}".format(e)
                list.append(x)
            elif e=='HDSP AL':
                x="{}".format(e)
                list.append(x)
            elif e=='AIR FLOW SWITCH ALARM':
                x="{}".format(e)
                list.append(x)
            elif e=='WATER FLOW SWITCH FAULT1':
                x="{}".format(e)
                list.append(x)
            elif e=='Digital Compressor OUTPUT':
                x="{}".format(e)
                list.append(x)
            elif e=='VFD Compressor OUTPUT':
                x="{}".format(e)
                list.append(x)
            elif e=='ew_vfd_out':
                x="{}".format(e)
                list.append(x)
            elif e=='Dehum Setpoint':
                x="{}".format(e)
                list.append(x)
                x="{}".format("Return Air Dehum Stpt UNOCC_Write")
                list.append(x)
            elif e=='hgrc Setpoint':
                x="{}".format(e)
                list.append(x)
                x="{}".format("HGR SETPT OFFSET_WRITE")
                list.append(x)
                x="{}".format("HGRV_OUT")
                list.append(x)
                x="{}".format("HGRV_OUT2")
                list.append(x)

            elif e=='Outside Air Humidity':
                x="{}".format(e)
                list.append(x)
            elif e=='sa_humidity_r':
                x="{}".format(e)
                list.append(x)
            elif e=='ra_humidity_r':
                x="{}".format(e)
                list.append(x)
            elif e=='Ea_humidity_r':
                x="{}".format(e)
                list.append(x)
            elif e=='ma_humidity_r':
                x="{}".format(e)
                list.append(x)
            elif e=='EWEA SENSOR':
                x="{}".format(e)
                list.append(x)
            elif e=='DXET SENSOR':
                x="{}".format(e)
                list.append(x)
            elif e=='RAT SENSOR':
                x="{}".format(e)
                list.append(x)
            elif e=='MAT SENSOR':
                x="{}".format(e)
                list.append(x)
            elif e=='OA DAMPER Output %':
                x="{}".format(e)
                list.append(x)
            elif e=='EA DAMPER Output %':
                x="{}".format(e)
                list.append(x)
            elif e=='RA DAMPER Output %':
                x="{}".format(e)
                list.append(x)
            elif e=='FRESH Damper1 S/S Status':
                x="{}".format(e)
                list.append(x)
            elif e=='Duct Pressure Reading':
                x="{}".format(e)
                list.append(x)
                x="{}".format("Duct Static Pressure Setpt")
                list.append(x)
            elif e=='Bldg Static Pressure RDG':
                x="{}".format(e)
                list.append(x)
                x="{}".format("Bldg Static Pressure Setpt")
                list.append(x)
            elif e=='SR1_Failure_Alarm':
              for i in range(1,(int(SRQTY)+1)):
                x="{}{}{}{}".format("SR",i," ","Failure Alarm")
                list.append(x)
                x="{}{}{}{}".format("CFM",i," ","Failure Alarm")
                list.append(x)
                x="{}{}{}{}".format("CFM",i," ","SPEED OUT")
                list.append(x)


        # print(list)
        # remove the temporary files if exists
        if os.path.exists("Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx"):
            os.remove("Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx")
        else:
            print("The file does not exist1")

        # remove the draft file if exists
        if os.path.exists("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx"):
            os.remove("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx")
        else:
            print("The file does not exist2")

        # print(list)
        wb=openpyxl.Workbook()
        wb['Sheet'].title="SHHEET"
        sheet = wb.active

        for v in range (0,len(list)):

            cv=sheet.cell(row=v+1,column=1)
            cv.value=list[v]
            # print(cv.value)

        
        wb.save("Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx")  # this is temp file contains the options names only
        # time.sleep(7)

        ################ we will bring the names from general sheet by comparing the above file with general sheet####################

        final_names1=read_xls1(general_list,"Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx").final_names
        # final_type1=read_xls1("X:\MALEK\Data\ALC\POINTS LISTS\general point list.xlsx",""X:/MALEK/Data/ALC/POINTS LISTS/temp file/11.xlsx").final_type
        # final_object_id1=read_xls1("X:\MALEK\Data\ALC\POINTS LISTS\general point list.xlsx",""X:/MALEK/Data/ALC/POINTS LISTS/temp file/11.xlsx").final_object_id
        # final_device_id1=read_xls1("X:\MALEK\Data\ALC\POINTS LISTS\general point list.xlsx",""X:/MALEK/Data/ALC/POINTS LISTS/temp file/11.xlsx").final_device_id
        # final_object_name1=read_xls1("X:\MALEK\Data\ALC\POINTS LISTS\general point list.xlsx",""X:/MALEK/Data/ALC/POINTS LISTS/temp file/11.xlsx").final_object_name
        final_read_write1=read_xls1(general_list,"Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx").final_read_write
        final_unit1=read_xls1(general_list,"Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx").final_unit
        final_min1=read_xls1(general_list,"Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx").final_min
        final_max1=read_xls1(general_list,"Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx").final_max
        final_normal_state1=read_xls1(general_list,"Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx").final_normal_state
        final_desc1=read_xls1(general_list,"Z:/M.hammad/PY/My_Apps/Excel files/temp file/temp file.xlsx").final_desc



        ##### below line will export a draft xls file which is similar to function pointlist1(the old one)###############
        write_xl1(final_names1,final_read_write1,final_unit1,final_min1,final_max1,final_normal_state1,final_desc1)

        ####### now we will get everything as final and ready to print it in final excel#########
        if general_list and alaa_list:
            print("eeeeeee")
           

            Project_Name=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).Project_Name
            Unit_Name=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).Unit_Name
            final_names2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_names1
            final_names2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_names1
            final_type2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_type1
            final_object_id2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_object_id1
            final_device_id2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_device_id1
            final_object_name2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_object_name1
            final_read_write2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_read_write1
            final_unit2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_unit1
            final_min2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_min1
            final_max2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_max1
            final_normal_state2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_normal_state1
            final_desc2=read_xls2("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/draft.xlsx",alaa_list).final_desc1a
            print(Project_Name)
            UN=Unit_Name.replace("-Main Program-1","")
            print(UN)

            write_excel(Project_Name,UN,final_names2,final_type2,final_object_id2,final_device_id2,final_object_name2,final_read_write2,final_unit2,final_min2,final_max2,final_normal_state2,final_desc2)
            # print(final_names2)
            # print(final_read_write1)
            # wb1=openpyxl.Workbook()
            # wb1['Sheet'].title="SHHEET"
            # sheet1 = wb1.active

            # for v1 in range (0,len(final_names1)):

            #     cv1=sheet1.cell(row=v1+1,column=1)
            #     cv1.value=final_names1[v1]
            # wb1.save("C:/Users/m-hamad/Desktop/me.xlsx")
            # print(final_names1)
            # print(Model)
            # print(DXQTY)
            # print(HEATERQTY)
            # # print(options)
            # print(list)
            # print(LOC)
            list.clear()
            # print(list)

            LOC="{}{}{}{}{}".format("Z:/M.hammad/PY/My_Apps/Excel files/output pointlist/",Project_Name," ",UN,".xlsx")
            # wb.save(LOC)

            ### DOWNLOAD THE FILE
            return send_file(LOC,as_attachment=True)  ### THIS SEND FILE ONLY WORKS OK WHEN IT'S USED IN VIEW FUNCTION, WHILE OTHER MODULES NOT SHOWING DOWNLOAD PROCESS
            # return render_template ("pointlist2.html",username=username)
       
    else:
        return render_template ("pointlist2.html",username=username)


   
