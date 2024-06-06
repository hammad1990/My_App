from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



### get data from master sheet
class read_excel:
    def __init__(self,path1,path2):
        self.path1=path1
        self.path2=path2
        workbook = load_workbook(filename=path1)
        workbook1 = load_workbook(filename=path2)
        self.sheet1 = workbook.active
        self.sheet2 = workbook1.active
###############################        
        self.final_names=[]
        self.final_type=[]
        self.final_object_id=[]
        self.final_device_id=[]
        self.final_object_name=[]
        self.final_read_write=[]
        self.final_unit=[]
        self.final_min=[]
        self.final_max=[]
        self.final_normal_state=[]
        self.final_desc=[]
###############################         
        
        self.mst_points=[]
        self.read_write=[]
        self.unit=[]
        self.min=[]
        self.max=[]
        self.normal_state=[]
        self.desc=[]
###############################
        self.pls_points=[]
        self.type=[]
        self.object_id=[]
        self.device_id=[]
        self.object_name=[]
        
        self.get_master()
        self.get_alaa()
    def get_master(self):   
        for i1 in range(2,self.sheet1.max_row+1):  ## start from row 2
            for j1 in range(3,4):   ##get point names
                cell1=self.sheet1.cell(row=i1,column=j1)
                self.mst_points.append(cell1.value)
            for j1h in range(8,9):   ### get Read/Write
                cell1h=self.sheet1.cell(row=i1,column=j1h)
                self.read_write.append(cell1h.value)
            for j1c in range(9,10):   ### get unit
                cell1c=self.sheet1.cell(row=i1,column=j1c)
                self.unit.append(cell1c.value)
            for j1a in range(10,11):   ### get min
                cell1a=self.sheet1.cell(row=i1,column=j1a)
                self.min.append(cell1a.value)
            for j1b in range(11,12):   ### get max
                cell1b=self.sheet1.cell(row=i1,column=j1b)
                self.max.append(cell1b.value)
            for j1i in range(12,13):   ### get Normal Status
                cell1i=self.sheet1.cell(row=i1,column=j1i)
                self.normal_state.append(cell1i.value)
            for j1d in range(13,14):   ### get desc
                cell1d=self.sheet1.cell(row=i1,column=j1d)
                self.desc.append(cell1d.value)
            
##        print(self.mst_points)            
      
        ### get data from point list sheet
    def get_alaa(self):
                
        for i2 in range(8,self.sheet2.max_row+1):## start from row 8
            for j2 in range(3,4):       ##get names
                cell2=self.sheet2.cell(row=i2,column=j2)
                if cell2.value==None:
                    pass
                else:
                    self.pls_points.append(cell2.value)
            for j2a in range(4,5):     ##get type
                cell2a=self.sheet2.cell(row=i2,column=j2a)
                if cell2a.value==None:
                    pass
                else:
                    self.type.append(cell2a.value)

            for j2b in range(5,6):     ##get Object ID
                cell2b=self.sheet2.cell(row=i2,column=j2b)
                if cell2b.value==None:
                    pass
                else:
                    self.object_id.append(cell2b.value)

            for j2c in range(6,7):     ##get device id
                cell2c=self.sheet2.cell(row=i2,column=j2c)
                if cell2c.value==None:
                    pass
                else:
                    self.device_id.append(cell2c.value)

            for j2d in range(7,8):     ##get Object Name
                cell2d=self.sheet2.cell(row=i2,column=j2d)
                if cell2d.value==None:
                    pass
                else:
                    self.object_name.append(cell2d.value)
                    
            
##        print(self.pls_points)
        self.final()
            
                       
    def final(self):       
                    
        for x in range (0,len(self.mst_points)):
            for y in range (0,len(self.pls_points)):
                if self.mst_points[x]==self.pls_points[y]:
                    self.final_names.append(self.mst_points[x])
                    self.final_type.append(self.type[y])
                    self.final_object_id.append(self.object_id[y])
                    self.final_device_id.append(self.device_id[y])
                    self.final_object_name.append(self.object_name[y])
                    self.final_read_write.append(self.read_write[x])
                    self.final_unit.append(self.unit[x])
                    self.final_min.append(self.min[x])
                    self.final_max.append(self.max[x])
                    self.final_normal_state.append(self.normal_state[x])
                    self.final_desc.append(self.desc[x])
                    
                    
                    break
##        print(self.final_desc,"\n")            
        count=0
        self.final_desc1=self.final_desc

           
        for d in list(self.final_desc1):

            
            
            count=0
            for e in list(self.final_desc):
                zz=self.final_desc.index(e)
                
                if d == e:
                    count +=1
            

                if count==2:
                    del self.final_names[zz]
                    del self.final_type[zz]
                    del self.final_object_id[zz]
                    del self.final_device_id[zz]
                    del self.final_object_name[zz]
                    del self.final_read_write[zz]
                    del self.final_unit[zz]
                    del self.final_min[zz]
                    del self.final_max[zz]
                    del self.normal_state[zz]
                    del self.final_desc[zz]
                    
                    
                    
                    
                    
                    count=0
                    break

##        print(self.normal_state,"\n") 


                    
##        for z in range (0,len(self.final_names)):
##            print(self.final_names[z]," ",self.final_type[z]," ",self.final_object_id[z]," ",self.final_device_id[z]," ",self.final_object_name[z]," ",self.final_read_write[z]," ",self.final_unit," ",
##                  self.final_min," ",self.final_max," ",self.normal_state," ",self.final_desc,"\n")

        return (self.final_names,self.final_type,self.final_object_id,self.final_device_id,self.final_object_name,self.final_read_write,self.final_unit,self.final_min,self.final_max,self.final_normal_state,self.final_desc)
        

if __name__=='__main__':
    p1="general point list.xlsx"
    p2="P212556-morgan_lewis_-_2222_market_st_point list.xlsx"
    rtx=read_excel(p1,p2)
    



