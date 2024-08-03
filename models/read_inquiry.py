from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlrd
import pathlib

### get data from inquiry sheet
class read_inquiry:
    def __init__(self,path):
        # print(path)
        
        self.path=path
        # print(self.path)
        # self.name=pathlib.Path(self.path).suffix
        # self.name=str(self.path)

        self.No_of_items=[]
        self.petra_codes=[]
        self.desc=[]
        self.qty=[]
        self.unit=[]
        self.get_data()
        
    def get_data(self):
        # print("fdfdfd")

        if  self.path.endswith('.xlsx'):
       
            print("file ends with xlsx")
            workbook_xlsx = load_workbook(filename=self.path)
           
            # print(workbook_xlsx)
            # print("xlsx")
            self.sheet_xlsx = workbook_xlsx.active
            self.inquiry_number=self.sheet_xlsx.cell(row=1,column=1).value## get inquiry Number
            self.inquiry_number=self.inquiry_number.replace("INQUIRY MATERIAL-","")
            self.inquiry_number=self.inquiry_number.replace("-a1","")
            print(self.inquiry_number)
            # self.supplier=self.sheet.cell(row=3,column=1).value## get supplier name
            # self.supplier=self.supplier.replace("SUPPLIER :","")
            # self.supplier=self.supplier.strip()
            # print(self.supplier)
            for i1 in range(4,self.sheet_xlsx.max_row+1):  ## start from row 5
                for j1 in range(1,2):   ##get number of items
                    cell1=self.sheet_xlsx.cell(row=i1,column=j1)
                    self.No_of_items.append(cell1.value)
                for j1h in range(2,3):   ### get petra codes
                    cell1h=self.sheet_xlsx.cell(row=i1,column=j1h)
                    self.petra_codes.append(cell1h.value)
                for j1c in range(4,5):   ### get desc
                    cell1c=self.sheet_xlsx.cell(row=i1,column=j1c)
                    self.desc.append(cell1c.value)
                for j1f in range(5,6):   ### get qty
                    cell1f=self.sheet_xlsx.cell(row=i1,column=j1f)
                    self.qty.append(cell1f.value)
                for j1w in range(6,7):   ### get unit
                    cell1w=self.sheet_xlsx.cell(row=i1,column=j1w)
                    self.unit.append(cell1w.value)

            # print(self.No_of_items[0])
            # print(self.petra_codes[0])
            # print(self.desc[0])
            # print(self.qty[0])
            # print(self.unit[0])

            return (self.inquiry_number,self.No_of_items,self.petra_codes,self.desc,self.qty,self.unit)
        
        elif self.path.endswith('.xls'):
            print("file ends with xls")
            workbook_xls = xlrd.open_workbook(filename=self.path)
            
            # print(workbook_xls)
            # print("The number of worksheets is {0}".format(workbook_xls.nsheets))
            self.sheet_xls = workbook_xls.sheet_by_index(0)
            self.inquiry_number=self.sheet_xls.cell(rowx=0,colx=0).value## get inquiry Number
            self.inquiry_number=self.inquiry_number.replace("INQUIRY MATERIAL-","")
            self.inquiry_number=self.inquiry_number.replace("-a1","")
            self.inquiry_number=self.inquiry_number.replace("-a2","")
            # print(self.inquiry_number)
            # self.supplier=self.sheet_xls.cell(row=3,column=1).value## get supplier name
            # self.supplier=self.supplier.replace("SUPPLIER :","")
            # self.supplier=self.supplier.strip()
            # print(self.supplier)
            # print(self.sheet_xls.nrows)
            for i1 in range(4,self.sheet_xls.nrows):  ## start from row 5
                for j1 in range(0,1):   ##get number of items
                    cell1=self.sheet_xls.cell(rowx=i1,colx=j1).value
                    self.No_of_items.append(cell1)
                    # print(cell1)
                for j1h in range(1,2):   ### get petra codes
                    cell1h=self.sheet_xls.cell(rowx=i1,colx=j1h).value
                    self.petra_codes.append(cell1h)
                for j1c in range(3,4):   ### get desc
                    cell1c=self.sheet_xls.cell(rowx=i1,colx=j1c).value
                    # cell1c=cell1c.encode('latin-1', 'replace').decode('latin-1') # this is to avoid using non latin-1 char since FPDF cannot accept it.
                    # cell1c=cell1c.replace("–","-")
                    # cell1c=cell1c.replace("”","(inch)")
                    # cell1c=cell1c.replace("—","-")
                    self.desc.append(cell1c)
                for j1f in range(4,5):   ### get qty
                    cell1f=self.sheet_xls.cell(rowx=i1,colx=j1f).value
                    self.qty.append(cell1f)
                    # print(cell1f)
                for j1w in range(5,6):   ### get unit
                    cell1w=self.sheet_xls.cell(rowx=i1,colx=j1w).value
                    self.unit.append(cell1w)

            # print(self.No_of_items[0])
            # print(self.petra_codes[0])
            # print(self.desc[0])
            # print(self.qty[0])
            # print(self.unit[0])

            return (self.inquiry_number,self.No_of_items,self.petra_codes,self.desc,self.qty,self.unit)


        


if __name__=='__main__':
    path="Z:\M.hammad\PY\My_Apps\INQUIRY MATERIAL-019905-a1.xls"
    path="Z:\\M.hammad\PY\My_Apps\INQUIRY MATERIAL-019755 DTECH.xlsx"
    rtx=read_inquiry(path)
    