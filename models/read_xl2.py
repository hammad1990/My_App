from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



### we will 
class read_xls2:
    def __init__(self,path1,path2):
        self.path1=path1
        self.path2=path2
        workbook1 = load_workbook(filename=path1)
        workbook2 = load_workbook(filename=path2)
        self.sheet1 = workbook1.active
        self.sheet2 = workbook2.active
        # print("dasda")
###############################  
        self.Project_Name=""  
        self.Unit_Name=""     
        self.final_names1=[]
        self.final_type1=[]
        self.final_object_id1=[]
        self.final_device_id1=[]
        self.final_object_name1=[]
        self.final_read_write1=[]
        self.final_unit1=[]
        self.final_min1=[]
        self.final_max1=[]
        self.final_normal_state1=[]
        self.final_desc1a=[]
###############################         
        
        self.mst_points=[]
        self.read_write=[]
        self.unit=[]
        self.min=[]
        self.max=[]
        self.normal_state=[]
        self.desc=[]
###############################
        self.pls_points=[]
        self.type=[]
        self.object_id=[]
        self.device_id=[]
        self.object_name=[]
        
        self.get_out1()
        self.get_alaa2()
    def get_out1(self):   
        for i1 in range(2,self.sheet1.max_row+1):  ## start from row 2
            for j1 in range(3,4):   ##get point names
                cell1=self.sheet1.cell(row=i1,column=j1)
                self.mst_points.append(cell1.value)
            for j1h in range(8,9):   ### get Read/Write
                cell1h=self.sheet1.cell(row=i1,column=j1h)
                self.read_write.append(cell1h.value)
            for j1c in range(9,10):   ### get unit
                cell1c=self.sheet1.cell(row=i1,column=j1c)
                self.unit.append(cell1c.value)
            for j1a in range(10,11):   ### get min
                cell1a=self.sheet1.cell(row=i1,column=j1a)
                self.min.append(cell1a.value)
            for j1b in range(11,12):   ### get max
                cell1b=self.sheet1.cell(row=i1,column=j1b)
                self.max.append(cell1b.value)
            for j1i in range(12,13):   ### get Normal Status
                cell1i=self.sheet1.cell(row=i1,column=j1i)
                self.normal_state.append(cell1i.value)
            for j1d in range(13,14):   ### get desc
                cell1d=self.sheet1.cell(row=i1,column=j1d)
                self.desc.append(cell1d.value)
            
        # print(self.mst_points)            
      
        ### get data from point list sheet
    def get_alaa2(self):
        self.Project_Name=self.sheet2.cell(2,2)   ##GET project name
        self.Unit_Name=self.sheet2.cell(7,2)      ##GET unit name
        self.Project_Name=self.Project_Name.value 
        self.Unit_Name=self.Unit_Name.value     


        for i2 in range(8,self.sheet2.max_row+1):## start from row 1
            for j2 in range(3,4):       ##get names
                cell2=self.sheet2.cell(row=i2,column=j2)
                if cell2.value==None:
                    # print("11")
                    pass
                else:
                    self.pls_points.append(cell2.value)
                    # print("22")
            for j2a in range(4,5):     ##get type
                cell2a=self.sheet2.cell(row=i2,column=j2a)
                if cell2a.value==None:
                    pass
                else:
                    self.type.append(cell2a.value)

            for j2b in range(5,6):     ##get Object ID
                cell2b=self.sheet2.cell(row=i2,column=j2b)
                if cell2b.value==None:
                    pass
                else:
                    self.object_id.append(cell2b.value)

            for j2c in range(6,7):     ##get device id
                cell2c=self.sheet2.cell(row=i2,column=j2c)
                if cell2c.value==None:
                    pass
                else:
                    self.device_id.append(cell2c.value)

            for j2d in range(7,8):     ##get Object Name
                cell2d=self.sheet2.cell(row=i2,column=j2d)
                if cell2d.value==None:
                    pass
                else:
                    self.object_name.append(cell2d.value)
                    
            
        # print(self.pls_points)
        self.final2()
            
                       
    def final2(self):       
                    
        for x in range (0,len(self.mst_points)):
            for y in range (0,len(self.pls_points)):
                if self.mst_points[x]==self.pls_points[y]:
                    self.final_names1.append(self.mst_points[x])
                    self.final_type1.append(self.type[y])
                    self.final_object_id1.append(self.object_id[y])
                    self.final_device_id1.append(self.device_id[y])
                    self.final_object_name1.append(self.object_name[y])
                    self.final_read_write1.append(self.read_write[x])
                    self.final_unit1.append(self.unit[x])
                    self.final_min1.append(self.min[x])
                    self.final_max1.append(self.max[x])
                    self.final_normal_state1.append(self.normal_state[x])
                    self.final_desc1a.append(self.desc[x])
                    
                    
                    break
##        print(self.final_desc,"\n")            
        count=0
        self.final_desc2=self.final_desc1a

           
        for d in list(self.final_desc2):

            
            
            count=0
            for e in list(self.final_desc1a):
                zz=self.final_desc1a.index(e)
                
                if d == e:
                    count +=1
            

                if count==2:
                    del self.final_names1[zz]
                    del self.final_type1[zz]
                    del self.final_object_id1[zz]
                    del self.final_device_id1[zz]
                    del self.final_object_name1[zz]
                    del self.final_read_write1[zz]
                    del self.final_unit1[zz]
                    del self.final_min1[zz]
                    del self.final_max1[zz]
                    del self.final_normal_state1[zz]
                    del self.final_desc1a[zz]
                    
                    
                    
                    
                    
                    count=0
                    break

##        print(self.normal_state,"\n") 


                    
##        for z in range (0,len(self.final_names)):
##            print(self.final_names[z]," ",self.final_type[z]," ",self.final_object_id[z]," ",self.final_device_id[z]," ",self.final_object_name[z]," ",self.final_read_write[z]," ",self.final_unit," ",
##                  self.final_min," ",self.final_max," ",self.normal_state," ",self.final_desc,"\n")

        return (self.Project_Name,self.Unit_Name,self.final_names1,self.final_type1,self.final_object_id1,self.final_device_id1,self.final_object_name1,self.final_read_write1,self.final_unit1,self.final_min1,self.final_max1,self.final_normal_state1,self.final_desc1a)
        # return (self.final_names)
        # return (self.final_names,self.final_read_write,self.final_unit,self.final_min,self.final_max,self.final_normal_state,self.final_desc)

if __name__=='__main__':
    p1="general point list.xlsx"
    p2="P212556-morgan_lewis_-_2222_market_st_point list.xlsx"
    rtx=read_xls2(p1,p2)
    



